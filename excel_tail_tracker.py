from __future__ import annotations

from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook, load_workbook


class ExcelTailTracker:
    def __init__(
        self,
        workbook_path: str,
        starting_bankroll: float,
        base_risk_pct: float,
        min_multiplier: float,
        max_multiplier: float,
        max_trade_notional_pct: float,
        max_market_notional_pct: float,
        max_account_notional_pct: float,
        multiplier_curve_power: float = 1.35,
        low_size_threshold_ratio: float = 0.12,
        low_size_haircut_power: float = 0.5,
        low_size_haircut_min_factor: float = 0.35,
    ):
        self.workbook_path = Path(workbook_path)
        self.run_start_utc = datetime.now(timezone.utc).isoformat()
        self.starting_bankroll = max(1.0, float(starting_bankroll))
        self.base_risk_pct = max(0.0, float(base_risk_pct))
        self.min_multiplier = float(min_multiplier)
        self.max_multiplier = float(max_multiplier)
        self.max_trade_notional_pct = max(0.0, float(max_trade_notional_pct))
        self.max_market_notional_pct = max(0.0, float(max_market_notional_pct))
        self.max_account_notional_pct = max(0.0, float(max_account_notional_pct))
        self.multiplier_curve_power = max(0.01, float(multiplier_curve_power))
        self.low_size_threshold_ratio = max(0.0, float(low_size_threshold_ratio))
        self.low_size_haircut_power = max(0.01, float(low_size_haircut_power))
        self.low_size_haircut_min_factor = max(0.0, min(1.0, float(low_size_haircut_min_factor)))

        self.realized_pnl = 0.0
        self.realized_gains = 0.0
        self.realized_losses = 0.0
        self.processed_trade_count = 0

        self.size_history_by_wallet: Dict[str, List[float]] = defaultdict(list)
        self.positions: Dict[str, Dict[str, Any]] = {}
        self.account_open_notional: Dict[str, float] = defaultdict(float)
        self.market_open_notional: Dict[str, float] = defaultdict(float)
        self.account_stats: Dict[str, Dict[str, float]] = defaultdict(
            lambda: {
                "copied_trades": 0,
                "copied_notional": 0.0,
                "realized_pnl": 0.0,
                "realized_gains": 0.0,
                "realized_losses": 0.0,
                "open_notional": 0.0,
            }
        )

        self._initialize_workbook()

    def set_wallet_size_history(self, wallet: str, observed_sizes: List[float]) -> None:
        sanitized: List[float] = []
        for value in observed_sizes:
            numeric = self._to_float(value, default=0.0)
            if numeric > 0:
                sanitized.append(numeric)

        self.size_history_by_wallet[wallet] = sanitized

    def wallet_has_open_positions(self, wallet: str) -> bool:
        wallet_prefix = f"{wallet}|"
        for position_key, position in self.positions.items():
            if position_key.startswith(wallet_prefix) and position.get("shares", 0.0) > 0:
                return True
        return False

    def wallets_with_open_positions(self) -> Set[str]:
        wallets: Set[str] = set()
        for position_key, position in self.positions.items():
            if position.get("shares", 0.0) <= 0:
                continue
            wallet = str(position_key).split("|", 1)[0]
            wallets.add(wallet)
        return wallets

    def _initialize_workbook(self) -> None:
        workbook = Workbook()
        trades = workbook.active
        trades.title = "Trades"
        trades.append(
            [
                "logged_at_utc",
                "trade_timestamp_utc",
                "source_wallet",
                "source_name",
                "market",
                "category",
                "outcome",
                "side",
                "observed_size",
                "observed_price",
                "size_percentile",
                "multiplier",
                "target_notional",
                "copied_notional",
                "copied_shares",
                "status",
                "realized_pnl_trade",
                "realized_pnl_total",
                "realized_roi_pct",
                "unsold_shares_value",
                "total_equity_est",
                "trade_key",
            ]
        )

        summary = workbook.create_sheet("Summary")
        summary.append(
            [
                "run_start_utc",
                "last_update_utc",
                "starting_bankroll",
                "realized_pnl",
                "realized_gains",
                "realized_losses",
                "realized_roi_pct",
                "unsold_shares_value",
                "total_equity_est",
                "open_positions",
                "tracked_accounts",
                "processed_trades",
            ]
        )
        summary.append([
            self.run_start_utc,
            self.run_start_utc,
            self.starting_bankroll,
            0.0,
            0.0,
            0.0,
            0.0,
            0.0,
            self.starting_bankroll,
            0,
            0,
            0,
        ])

        positions = workbook.create_sheet("Positions")
        positions.append(
            [
                "source_wallet",
                "source_name",
                "market",
                "outcome",
                "shares",
                "avg_cost",
                "last_price",
                "current_value",
                "cost_basis",
                "unrealized_pnl",
            ]
        )

        accounts = workbook.create_sheet("Accounts")
        accounts.append(
            [
                "source_wallet",
                "source_name",
                "copied_trades",
                "copied_notional",
                "realized_pnl",
                "realized_gains",
                "realized_losses",
                "open_notional",
            ]
        )

        workbook.save(self.workbook_path)

    @staticmethod
    def _to_float(value: Any, default: float = 0.0) -> float:
        try:
            return float(value)
        except (TypeError, ValueError):
            return default

    @staticmethod
    def _to_timestamp_iso(value: Any) -> str:
        try:
            ts = float(value)
            if ts > 1_000_000_000_000:
                ts /= 1000.0
            return datetime.fromtimestamp(ts, tz=timezone.utc).isoformat()
        except (TypeError, ValueError, OSError):
            return ""

    def _size_percentile(self, wallet: str, observed_size: float) -> float:
        history = self.size_history_by_wallet[wallet]
        if not history:
            history.append(observed_size)
            return 0.5

        count_leq = sum(1 for item in history if item <= observed_size)
        percentile = count_leq / len(history)
        history.append(observed_size)
        return max(0.0, min(1.0, percentile))

    @staticmethod
    def _median(values: List[float]) -> float:
        if not values:
            return 0.0
        ordered = sorted(values)
        size = len(ordered)
        midpoint = size // 2
        if size % 2 == 1:
            return ordered[midpoint]
        return (ordered[midpoint - 1] + ordered[midpoint]) / 2.0

    def _hybrid_multiplier(self, percentile: float, observed_size: float, wallet_median_size: float) -> float:
        percentile = max(0.0, min(1.0, percentile))
        base_multiplier = self.min_multiplier + (self.max_multiplier - self.min_multiplier) * (percentile ** self.multiplier_curve_power)

        adjusted_multiplier = base_multiplier
        if wallet_median_size > 0 and observed_size > 0 and self.low_size_threshold_ratio > 0:
            relative_size = observed_size / wallet_median_size
            if relative_size < self.low_size_threshold_ratio:
                haircut = max(self.low_size_haircut_min_factor, relative_size ** self.low_size_haircut_power)
                adjusted_multiplier *= haircut

        max_allowed = max(self.max_multiplier, self.min_multiplier)
        return max(0.0, min(adjusted_multiplier, max_allowed))

    @staticmethod
    def _instrument_parts(trade: Dict[str, Any]) -> Dict[str, str]:
        market = str(trade.get("title") or trade.get("slug") or trade.get("eventSlug") or "Unknown Market")
        outcome = str(trade.get("outcome") or "Unknown")
        base_id = str(trade.get("asset") or trade.get("slug") or trade.get("eventSlug") or trade.get("conditionId") or "unknown")
        market_key = f"{base_id}|{outcome}".lower()
        return {
            "market": market,
            "outcome": outcome,
            "market_key": market_key,
        }

    def _current_unsold_value(self) -> float:
        return sum(position["shares"] * position["last_price"] for position in self.positions.values())

    def _current_open_cost_basis(self) -> float:
        return sum(position["shares"] * position["avg_cost"] for position in self.positions.values())

    def _current_unrealized_pnl(self) -> float:
        return self._current_unsold_value() - self._current_open_cost_basis()

    def _current_free_bankroll(self) -> float:
        return self.starting_bankroll + self.realized_pnl - self._current_open_cost_basis()

    def _save_sheets(self) -> None:
        workbook = load_workbook(self.workbook_path)

        summary = workbook["Summary"]
        summary.cell(row=2, column=1, value=self.run_start_utc)
        summary.cell(row=2, column=2, value=datetime.now(timezone.utc).isoformat())
        summary.cell(row=2, column=3, value=self.starting_bankroll)
        summary.cell(row=2, column=4, value=self.realized_pnl)
        summary.cell(row=2, column=5, value=self.realized_gains)
        summary.cell(row=2, column=6, value=self.realized_losses)
        summary.cell(row=2, column=7, value=(self.realized_pnl / self.starting_bankroll) * 100.0)
        summary.cell(row=2, column=8, value=self._current_unsold_value())
        summary.cell(
            row=2,
            column=9,
            value=self.starting_bankroll + self.realized_pnl + self._current_unrealized_pnl(),
        )
        summary.cell(row=2, column=10, value=len(self.positions))
        summary.cell(row=2, column=11, value=len(self.account_stats))
        summary.cell(row=2, column=12, value=self.processed_trade_count)

        positions_sheet = workbook["Positions"]
        if positions_sheet.max_row > 1:
            positions_sheet.delete_rows(2, positions_sheet.max_row - 1)
        for position in self.positions.values():
            shares = position["shares"]
            avg_cost = position["avg_cost"]
            last_price = position["last_price"]
            current_value = shares * last_price
            cost_basis = shares * avg_cost
            positions_sheet.append(
                [
                    position["wallet"],
                    position["display_name"],
                    position["market"],
                    position["outcome"],
                    shares,
                    avg_cost,
                    last_price,
                    current_value,
                    cost_basis,
                    current_value - cost_basis,
                ]
            )

        accounts_sheet = workbook["Accounts"]
        if accounts_sheet.max_row > 1:
            accounts_sheet.delete_rows(2, accounts_sheet.max_row - 1)
        for wallet, stats in self.account_stats.items():
            accounts_sheet.append(
                [
                    wallet,
                    stats.get("display_name") or wallet,
                    stats["copied_trades"],
                    stats["copied_notional"],
                    stats["realized_pnl"],
                    stats["realized_gains"],
                    stats["realized_losses"],
                    self.account_open_notional.get(wallet, 0.0),
                ]
            )

        workbook.save(self.workbook_path)

    def record_trade(self, trade: Dict[str, Any], user: Dict[str, Any], category: str, trade_key: str) -> None:
        wallet = str(user.get("wallet") or "unknown_wallet")
        display_name = str(user.get("display_name") or wallet)
        side = str(trade.get("side") or "").upper()
        observed_size = max(0.0, self._to_float(trade.get("size"), default=0.0))
        observed_price = max(0.0, min(1.0, self._to_float(trade.get("price"), default=0.0)))

        parts = self._instrument_parts(trade)
        market = parts["market"]
        outcome = parts["outcome"]
        market_key = parts["market_key"]
        position_key = f"{wallet}|{market_key}"

        wallet_median_size = self._median(self.size_history_by_wallet.get(wallet, []))
        percentile = self._size_percentile(wallet, observed_size)
        multiplier = self._hybrid_multiplier(
            percentile=percentile,
            observed_size=observed_size,
            wallet_median_size=wallet_median_size,
        )

        target_notional = self.starting_bankroll * self.base_risk_pct * multiplier
        per_trade_cap = self.starting_bankroll * self.max_trade_notional_pct

        max_account_open = self.starting_bankroll * self.max_account_notional_pct
        max_market_open = self.starting_bankroll * self.max_market_notional_pct
        remaining_account = max(0.0, max_account_open - self.account_open_notional.get(wallet, 0.0))
        remaining_market = max(0.0, max_market_open - self.market_open_notional.get(market_key, 0.0))
        remaining_cash = max(0.0, self._current_free_bankroll())

        copied_notional = 0.0
        copied_shares = 0.0
        realized_trade_pnl = 0.0
        status = "IGNORED"

        account_stats = self.account_stats[wallet]
        account_stats["display_name"] = display_name

        if side == "BUY" and observed_price > 0:
            copied_notional = min(target_notional, per_trade_cap, remaining_account, remaining_market, remaining_cash)
            if copied_notional > 0:
                copied_shares = copied_notional / observed_price
                position = self.positions.get(position_key)
                if not position:
                    position = {
                        "wallet": wallet,
                        "display_name": display_name,
                        "market": market,
                        "outcome": outcome,
                        "market_key": market_key,
                        "shares": 0.0,
                        "avg_cost": 0.0,
                        "last_price": observed_price,
                    }
                    self.positions[position_key] = position

                old_shares = position["shares"]
                new_shares = old_shares + copied_shares
                if new_shares > 0:
                    position["avg_cost"] = ((old_shares * position["avg_cost"]) + (copied_shares * observed_price)) / new_shares
                position["shares"] = new_shares
                position["last_price"] = observed_price

                self.account_open_notional[wallet] = self.account_open_notional.get(wallet, 0.0) + copied_notional
                self.market_open_notional[market_key] = self.market_open_notional.get(market_key, 0.0) + copied_notional
                status = "OPENED"

        elif side == "SELL" and observed_price > 0:
            position = self.positions.get(position_key)
            if position and position["shares"] > 0:
                if observed_price >= 1.0 - 1e-9:
                    close_shares = position["shares"]
                else:
                    close_target_notional = min(target_notional, per_trade_cap)
                    close_shares = min(position["shares"], close_target_notional / observed_price)
                if close_shares > 0:
                    copied_shares = close_shares
                    copied_notional = close_shares * observed_price
                    realized_trade_pnl = (observed_price - position["avg_cost"]) * close_shares

                    position["shares"] -= close_shares
                    position["last_price"] = observed_price

                    released_cost_basis = close_shares * position["avg_cost"]
                    self.account_open_notional[wallet] = max(0.0, self.account_open_notional.get(wallet, 0.0) - released_cost_basis)
                    self.market_open_notional[market_key] = max(
                        0.0,
                        self.market_open_notional.get(market_key, 0.0) - released_cost_basis,
                    )

                    if position["shares"] <= 1e-10:
                        del self.positions[position_key]

                    status = "CLOSED"

        if status != "IGNORED":
            self.processed_trade_count += 1
            self.realized_pnl += realized_trade_pnl
            if realized_trade_pnl >= 0:
                self.realized_gains += realized_trade_pnl
                account_stats["realized_gains"] += realized_trade_pnl
            else:
                self.realized_losses += abs(realized_trade_pnl)
                account_stats["realized_losses"] += abs(realized_trade_pnl)

            account_stats["copied_trades"] += 1
            account_stats["copied_notional"] += copied_notional
            account_stats["realized_pnl"] += realized_trade_pnl

        unsold_value = self._current_unsold_value()
        unrealized_pnl = self._current_unrealized_pnl()
        realized_roi_pct = (self.realized_pnl / self.starting_bankroll) * 100.0
        total_equity_est = self.starting_bankroll + self.realized_pnl + unrealized_pnl

        workbook = load_workbook(self.workbook_path)
        trades = workbook["Trades"]
        trades.append(
            [
                datetime.now(timezone.utc).isoformat(),
                self._to_timestamp_iso(trade.get("timestamp")),
                wallet,
                display_name,
                market,
                category,
                outcome,
                side,
                observed_size,
                observed_price,
                percentile,
                multiplier,
                target_notional,
                copied_notional,
                copied_shares,
                status,
                realized_trade_pnl,
                self.realized_pnl,
                realized_roi_pct,
                unsold_value,
                total_equity_est,
                trade_key,
            ]
        )
        workbook.save(self.workbook_path)

        self._save_sheets()
