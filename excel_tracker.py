"""
Excel-based performance tracking.

Creates and maintains an Excel workbook with:
- Balance history (time-series tracking)
- Trade log (all buys and sells)
- Position summary (current open positions)
"""
import logging
import os
from datetime import datetime, timezone
from typing import Any, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelTracker:
    """Excel workbook tracker for performance monitoring."""
    
    def __init__(self, filepath: str = "live_performance.xlsx"):
        """
        Initialize Excel tracker.
        
        Args:
            filepath: Path to Excel file
        """
        self.filepath = filepath
        self.wb: Optional[Workbook] = None
        self._initialize_workbook()
    
    def _initialize_workbook(self):
        """Initialize or load Excel workbook."""
        if os.path.exists(self.filepath):
            try:
                self.wb = load_workbook(self.filepath)
                logger.info(f"Loaded existing Excel tracker: {self.filepath}")
                
                # Ensure all required sheets exist
                self._ensure_sheets()
            except Exception as e:
                logger.exception(f"Error loading Excel file: {e}")
                self._create_new_workbook()
        else:
            self._create_new_workbook()
    
    def _create_new_workbook(self):
        """Create a new Excel workbook with required sheets."""
        self.wb = Workbook()
        
        # Remove default sheet
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]
        
        # Create sheets
        self._create_balance_sheet()
        self._create_trades_sheet()
        self._create_positions_sheet()
        
        self._save()
        logger.info(f"Created new Excel tracker: {self.filepath}")
    
    def _ensure_sheets(self):
        """Ensure all required sheets exist."""
        if "Balance History" not in self.wb.sheetnames:
            self._create_balance_sheet()
        
        if "Trades" not in self.wb.sheetnames:
            self._create_trades_sheet()
        
        if "Open Positions" not in self.wb.sheetnames:
            self._create_positions_sheet()
    
    def _create_balance_sheet(self):
        """Create Balance History sheet."""
        ws = self.wb.create_sheet("Balance History")
        
        # Header row
        headers = ["Timestamp", "Balance", "Invested", "Available", "Total Positions", "P&L"]
        ws.append(headers)
        
        # Format header
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
        
        # Set column widths
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 12
    
    def _create_trades_sheet(self):
        """Create Trades log sheet."""
        ws = self.wb.create_sheet("Trades")
        
        # Header row
        headers = [
            "Timestamp",
            "Market",
            "Outcome",
            "Side",
            "Shares",
            "Price",
            "Amount",
            "Trader",
            "Status",
        ]
        ws.append(headers)
        
        # Format header
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
        
        # Set column widths
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 8
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 10
        ws.column_dimensions["G"].width = 12
        ws.column_dimensions["H"].width = 15
        ws.column_dimensions["I"].width = 12
    
    def _create_positions_sheet(self):
        """Create Open Positions summary sheet."""
        ws = self.wb.create_sheet("Open Positions")
        
        # Header row
        headers = [
            "Market",
            "Outcome",
            "Shares",
            "Entry Price",
            "Invested",
            "Current Price",
            "Current Value",
            "P&L",
            "P&L %",
            "Opened",
            "Trader",
        ]
        ws.append(headers)
        
        # Format header
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
        
        # Set column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 12
        ws.column_dimensions["H"].width = 10
        ws.column_dimensions["I"].width = 10
        ws.column_dimensions["J"].width = 20
        ws.column_dimensions["K"].width = 15
    
    def _save(self):
        """Save workbook to file."""
        try:
            self.wb.save(self.filepath)
        except Exception as e:
            logger.exception(f"Error saving Excel file: {e}")
    
    def log_balance(
        self,
        balance: float,
        invested: float,
        total_positions: int,
        pnl: Optional[float] = None,
    ):
        """
        Log balance snapshot to Balance History sheet.
        
        Args:
            balance: Current balance
            invested: Total invested amount
            total_positions: Number of open positions
            pnl: Total P&L (optional)
        """
        try:
            ws = self.wb["Balance History"]
            
            timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
            available = balance - invested
            
            row = [
                timestamp,
                f"{balance:.2f}",
                f"{invested:.2f}",
                f"{available:.2f}",
                total_positions,
                f"{pnl:.2f}" if pnl is not None else "N/A",
            ]
            
            ws.append(row)
            self._save()
            
        except Exception as e:
            logger.exception(f"Error logging balance: {e}")
    
    def log_trade(
        self,
        market_slug: str,
        outcome: str,
        side: str,
        shares: float,
        price: float,
        trader: Optional[str] = None,
        status: str = "pending",
    ):
        """
        Log a trade to Trades sheet.
        
        Args:
            market_slug: Market slug
            outcome: Outcome
            side: "BUY" or "SELL"
            shares: Number of shares
            price: Trade price
            trader: Monitored trader (optional)
            status: Trade status
        """
        try:
            ws = self.wb["Trades"]
            
            timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
            amount = shares * price
            
            row = [
                timestamp,
                market_slug,
                outcome.upper(),
                side.upper(),
                f"{shares:.2f}",
                f"{price:.4f}",
                f"{amount:.2f}",
                trader or "N/A",
                status,
            ]
            
            ws.append(row)
            self._save()
            
        except Exception as e:
            logger.exception(f"Error logging trade: {e}")
    
    def update_positions(
        self,
        positions: list[dict[str, Any]],
        position_pnl_map: Optional[dict[str, dict[str, float]]] = None,
    ):
        """
        Update Open Positions sheet with current positions.
        
        Args:
            positions: List of position dicts
            position_pnl_map: Dict mapping position keys to P&L data (optional)
        """
        try:
            ws = self.wb["Open Positions"]
            
            # Clear existing data (keep header)
            ws.delete_rows(2, ws.max_row)
            
            # Add position rows
            for position in positions:
                market_slug = position["market_slug"]
                outcome = position["outcome"]
                shares = position["shares"]
                entry_price = position["entry_price"]
                invested = position.get("invested", shares * entry_price)
                opened_at = position.get("opened_at", "N/A")
                trader = position.get("monitored_trader", "N/A")
                
                # Parse opened_at timestamp
                if isinstance(opened_at, str) and opened_at != "N/A":
                    try:
                        dt = datetime.fromisoformat(opened_at.replace("Z", "+00:00"))
                        opened_at = dt.strftime("%Y-%m-%d %H:%M")
                    except:
                        pass
                
                # Get P&L data if available
                current_price = "N/A"
                current_value = "N/A"
                pnl = "N/A"
                pnl_pct = "N/A"
                
                if position_pnl_map:
                    position_key = f"{market_slug}|{outcome.lower()}"
                    pnl_data = position_pnl_map.get(position_key)
                    
                    if pnl_data:
                        current_value_float = pnl_data.get("current_value")
                        if current_value_float is not None and shares > 0:
                            current_price = f"{current_value_float / shares:.4f}"
                            current_value = f"{current_value_float:.2f}"
                            pnl = f"{pnl_data.get('pnl', 0):.2f}"
                            pnl_pct = f"{pnl_data.get('pnl_pct', 0):.2f}%"
                
                row = [
                    market_slug,
                    outcome.upper(),
                    f"{shares:.2f}",
                    f"{entry_price:.4f}",
                    f"{invested:.2f}",
                    current_price,
                    current_value,
                    pnl,
                    pnl_pct,
                    opened_at,
                    trader,
                ]
                
                ws.append(row)
            
            self._save()
            
        except Exception as e:
            logger.exception(f"Error updating positions: {e}")
    
    def close(self):
        """Close workbook."""
        if self.wb:
            try:
                self._save()
                self.wb.close()
                logger.info("Excel tracker closed")
            except Exception as e:
                logger.exception(f"Error closing Excel tracker: {e}")
