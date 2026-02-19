import os
import time
from datetime import datetime
from openpyxl import load_workbook

FILES = [
    ("80", "tail_performance.xlsx"),
    ("75", "tail_performance_75.xlsx"),
]


def read_summary(path):
    if not os.path.exists(path):
        return None
    wb = load_workbook(path, data_only=True)
    if "Summary" not in wb.sheetnames:
        return None
    ws = wb["Summary"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    values = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
    return {str(h): v for h, v in zip(headers, values) if h is not None}

print("Live workbook watcher started (Ctrl+C to stop)", flush=True)
while True:
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print("=" * 72, flush=True)
    print(f"{now}", flush=True)
    for label, path in FILES:
        data = read_summary(path)
        if not data:
            print(f"PROFILE {label} | workbook missing/unreadable: {path}", flush=True)
            continue
        trades = int(float(data.get("processed_trades") or 0))
        realized = float(data.get("realized_pnl") or 0)
        equity = float(data.get("total_equity_est") or 0)
        updated = data.get("last_update_utc")
        print(
            f"PROFILE {label} | trades={trades} | realized_pnl={realized:.2f} | equity={equity:.2f} | last_update={updated}",
            flush=True,
        )
    time.sleep(10)
