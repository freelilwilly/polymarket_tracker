import argparse
import os
import time
from datetime import datetime

from openpyxl import load_workbook


def to_float(value, default=0.0):
    try:
        if value is None:
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def print_report(workbook: str, label: str, warn_roi: float):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    print("=" * 72, flush=True)
    print(f"HOURLY TAIL METRICS | PROFILE {label} | {timestamp}", flush=True)
    print("-" * 72, flush=True)

    if not os.path.exists(workbook):
        print(f"Workbook: {workbook}", flush=True)
        print("Status: workbook not found", flush=True)
        print("=" * 72, flush=True)
        return

    wb = load_workbook(workbook, data_only=True)
    if "Summary" not in wb.sheetnames:
        print(f"Workbook: {workbook}", flush=True)
        print("Status: Summary sheet missing", flush=True)
        print("=" * 72, flush=True)
        return

    ws = wb["Summary"]

    headers = [ws.cell(row=1, column=column).value for column in range(1, ws.max_column + 1)]
    values = [ws.cell(row=2, column=column).value for column in range(1, ws.max_column + 1)]
    summary = {
        str(header): value
        for header, value in zip(headers, values)
        if header is not None
    }

    run_start = summary.get("run_start_utc")
    last_update = summary.get("last_update_utc")
    start_bankroll = to_float(summary.get("starting_bankroll"), 0.0)
    trades = int(to_float(summary.get("processed_trades"), 0.0))
    realized_pnl = to_float(summary.get("realized_pnl"), 0.0)
    realized_gains = to_float(summary.get("realized_gains"), 0.0)
    realized_losses = to_float(summary.get("realized_losses"), 0.0)
    unsold_value = to_float(summary.get("unsold_shares_value"), 0.0)
    ending_equity = to_float(summary.get("total_equity_est"), 0.0)
    roi = to_float(summary.get("realized_roi_pct"), 0.0)
    equity_roi = ((ending_equity - start_bankroll) / start_bankroll) * 100.0 if start_bankroll > 0 else 0.0
    open_positions = int(to_float(summary.get("open_positions"), 0.0))
    tracked_accounts = int(to_float(summary.get("tracked_accounts"), 0.0))

    print(f"Workbook: {workbook}", flush=True)
    print(f"Run Start: {run_start}", flush=True)
    print(f"Last Update: {last_update}", flush=True)
    print(f"Trades: {trades}", flush=True)
    print(f"Starting Bankroll: {start_bankroll:.2f}", flush=True)
    print(f"Realized PnL: {realized_pnl:.2f}", flush=True)
    print(f"Realized Gains: {realized_gains:.2f}", flush=True)
    print(f"Realized Losses: {realized_losses:.2f}", flush=True)
    print(f"Unsold Value: {unsold_value:.2f}", flush=True)
    print(f"Open Positions: {open_positions}", flush=True)
    print(f"Tracked Accounts: {tracked_accounts}", flush=True)
    print(f"Ending Equity: {ending_equity:.2f}", flush=True)
    print(f"Realized ROI: {roi:.2f}%", flush=True)
    print(f"Equity ROI: {equity_roi:.2f}%", flush=True)

    if roi <= warn_roi:
        print(f"WARNING: ROI {roi:.2f}% <= {warn_roi:.2f}%", flush=True)

    print("=" * 72, flush=True)


def main():
    parser = argparse.ArgumentParser(description="Hourly workbook reporter")
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--label", required=True)
    parser.add_argument("--warn-roi", type=float, default=-25.0)
    parser.add_argument("--interval-seconds", type=int, default=3600)
    args = parser.parse_args()

    interval_seconds = max(60, args.interval_seconds)

    print(
        f"Starting hourly reporter for profile {args.label} (interval={interval_seconds}s)",
        flush=True,
    )

    while True:
        try:
            print_report(args.workbook, args.label, args.warn_roi)
        except Exception as exc:
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            print("=" * 72, flush=True)
            print(f"HOURLY TAIL METRICS | PROFILE {args.label} | {now}", flush=True)
            print("-" * 72, flush=True)
            print(f"Reporter error: {exc}", flush=True)
            print("=" * 72, flush=True)

        time.sleep(interval_seconds)


if __name__ == "__main__":
    main()
